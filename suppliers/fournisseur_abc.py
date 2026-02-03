from __future__ import annotations

import re



def map_custom_product_type(val: str) -> str:
    """Map supplier product types to Shopify Custom Product Type (case-insensitive)."""
    if val is None:
        return val
    if not isinstance(val, str):
        val = str(val)
    v = val.strip().lower()

    mapping = {
        "gilet": "Vests",
        "bibs": "Bib Shorts",
        "long bibs": "Bib Tights",
        "bidon": "Water Bottle",
        "baselayer": "Base Layer",
        # keep existing convention for tees
        "t-shirt": "T-Shirts",
        "t shirt": "T-Shirts",
        "tee": "T-Shirts",
        "tshirt": "T-Shirts",
    }

    # Prefer exact match; fallback to contains for common tee variants
    if v in mapping:
        return mapping[v]

    # contains-based only for a couple of safe cases
    if "long bibs" in v:
        return "Bib Tights"
    if v.startswith("bibs") or " bibs" in v:
        return "Bib Shorts"
    if "t-shirt" in v or "t shirt" in v or " tee" in f" {v}" or "tshirt" in v:
        return "T-Shirts"

    return val


SIZE_REGEX = re.compile(r"""(\s*[-/]?\s*(?:size\s*)?(?:xs|s|m|l|xl|xxl)\b)""", re.IGNORECASE)

def remove_size(text):
    if not isinstance(text, str):
        return text
    return re.sub(SIZE_REGEX, "", text).strip()




def _strip_size_tokens(s: str) -> str:
    """Remove size tokens (XS/S/M/L/XL/XXL...) from SEO fields without removing letters inside words."""
    if s is None:
        return s
    if not isinstance(s, str):
        s = str(s)

    out = s

    # Remove parenthesized sizes, e.g. "(M)" or "(size M)"
    out = re.sub(r"(?i)\(\s*(?:size\s*)?(?:xxs|xs|s|m|l|xl|xxl|xxxl)\s*\)", "", out)

    # Remove size tokens only when preceded by a separator or whitespace (so we don't touch words like "Studios")
    out = re.sub(r"(?i)(?:\s*[-/]\s*|\s+)(?:size\s*)?(?:xxs|xs|s|m|l|xl|xxl|xxxl)\b", "", out)

    # Cleanup leftover separators/spaces
    out = re.sub(r"\s{2,}", " ", out).strip()
    out = re.sub(r"\s+([,;:.])", r"\1", out)
    out = re.sub(r"[-/]\s*$", "", out).strip()
    out = re.sub(r"\s*-\s*$", "", out).strip()

    return out
def _scrub_nan_token_in_title(s: str) -> str:
    """Remove accidental 'nan' tokens that can appear when concatenating missing fields."""
    if s is None:
        return ""
    t = str(s).replace("\u00A0", " ").strip()
    # remove leading 'nan - ' patterns
    t = re.sub(r"(?i)^\s*nan\s*-\s*", "", t)
    # remove standalone ' nan ' tokens (rare) and clean double spaces
    t = re.sub(r"(?i)\bnan\b", "", t)
    t = re.sub(r"\s{2,}", " ", t).strip()
    # remove leftover leading/trailing hyphens
    t = re.sub(r"^\s*-\s*", "", t)
    t = re.sub(r"\s*-\s*$", "", t)
    return t


def _norm_upc(v) -> str:
    """Normalize UPC/Barcode: keep digits only, drop trailing .0 from numeric."""
    if v is None:
        return ""
    s = str(v).strip()
    # drop .0 for floats represented as '123.0'
    s = re.sub(r"\.0$", "", s)
    # keep digits only
    s = re.sub(r"\D", "", s)
    return s



def _colkey(c: str) -> str:
    """Normalize column name: lower, remove spaces/punct/parentheses for robust matching."""
    s = str(c or "").strip().lower()
    s = re.sub(r"[\s\-_/]+", "", s)
    s = s.replace("(", "").replace(")", "")
    return s

def _find_col(df_cols, candidates):
    """Return first column in df_cols matching any normalized candidate."""
    norm_map = {_colkey(c): c for c in df_cols}
    for cand in candidates:
        k = _colkey(cand)
        if k in norm_map:
            return norm_map[k]
    # also allow partial contains on normalized
    cols_norm = [(_colkey(c), c) for c in df_cols]
    for cand in candidates:
        ck = _colkey(cand)
        for cn, orig in cols_norm:
            if ck and ck in cn:
                return orig
    return None


def _header_has_cad(col_name: str) -> bool:
    return "cad" in str(col_name or "").lower()



import io
import re
import math
import unicodedata

import pandas as pd

def _read_supplier_csv(file_like, filename: str) -> pd.DataFrame:
    """Read supplier CSV robustly (encoding + delimiter).

    - Tries several encodings (utf-8-sig/utf-8/cp1252/latin1).
    - Tries common delimiters and also auto-detects (sep=None).
    - Keeps empty cells empty (avoid NaN) with keep_default_na=False.
    """
    encodings = ["utf-8-sig", "utf-8", "cp1252", "latin1"]

    # 1) Try auto delimiter detection first for each encoding
    last_err = None
    for enc in encodings:
        try:
            try:
                file_like.seek(0)
            except Exception:
                pass
            df = pd.read_csv(
                file_like,
                encoding=enc,
                sep=None,
                engine="python",
                dtype=str,
                keep_default_na=False,
                encoding_errors="replace",
            )
            return df
        except Exception as e:
            last_err = e

    # 2) Fallback: try explicit separators
    seps = [",", ";", "	", "|"]
    for enc in encodings:
        for sep in seps:
            try:
                try:
                    file_like.seek(0)
                except Exception:
                    pass
                df = pd.read_csv(
                    file_like,
                    encoding=enc,
                    sep=sep,
                    engine="python",
                    dtype=str,
                    keep_default_na=False,
                    encoding_errors="replace",
                )
                # If we parsed a single giant column, try other seps
                if df.shape[1] == 1 and sep != seps[-1]:
                    continue
                return df
            except Exception as e:
                last_err = e

    raise ValueError(
        f"Impossible de lire le CSV fournisseur ({filename}). Encodages testés: {encodings}. Dernière erreur: {last_err}"
    )



def _series_str_clean(s: pd.Series) -> pd.Series:
    """Convert a series to clean strings without 'nan'/'none' tokens."""
    s2 = s.fillna("").astype(str).replace({r"^\s*(nan|none)\s*$": ""}, regex=True)
    return s2

import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment

# Optional dependency: python-slugify
try:
    from slugify import slugify  # type: ignore
except Exception:
    def slugify(value: str) -> str:
        s = str(value or "").strip().lower()
        s = re.sub(r"[^a-z0-9]+", "-", s)
        s = re.sub(r"-{2,}", "-", s).strip("-")
        return s

def _build_existing_shopify_index(existing_shopify_xlsx_bytes: bytes | None):
    """Build matching indexes from an existing Shopify product export/list.

    Matching rule (as requested):
      * if SKU + UPC : key = SKU|UPC
      * else UPC : key = UPC
      * else Vendor + SKU : key = Vendor|SKU
      * else : no key (do not classify as existing)

    Also returns a set of existing handles (normalized).
    """
    handles_set: set[str] = set()
    key_sets = {
        "sku_upc": set(),      # (sku, upc)
        "upc": set(),          # (upc,)
        "vendor_sku": set(),   # (vendor, sku)
    }
    if not existing_shopify_xlsx_bytes:
        return handles_set, key_sets

    try:
        bio = io.BytesIO(existing_shopify_xlsx_bytes)
        df = pd.read_excel(bio)  # first sheet by default
    except Exception:
        return handles_set, key_sets

    cols_l = {str(c).strip().lower(): c for c in df.columns}
    handle_col = cols_l.get("handle")
    vendor_col = cols_l.get("vendor") or cols_l.get("brand")
    sku_col = cols_l.get("variant sku") or cols_l.get("sku")
    upc_col = cols_l.get("variant barcode") or cols_l.get("barcode") or cols_l.get("upc")

    for _, r in df.iterrows():
        h = _norm_handle(r.get(handle_col, "")) if handle_col else ""
        if h:
            handles_set.add(h)

        vendor = _norm(r.get(vendor_col, "")) if vendor_col else ""
        sku = _norm(r.get(sku_col, "")) if sku_col else ""
        upc = _norm_upc(r.get(upc_col, "")) if upc_col else ""

        # Priority order for keys
        if sku and upc:
            key_sets["sku_upc"].add((sku, upc))
        elif upc:
            key_sets["upc"].add((upc,))
        elif vendor and sku:
            key_sets["vendor_sku"].add((vendor, sku))

    return handles_set, key_sets


def _row_is_existing(vendor: str, sku: str, upc: str, key_sets) -> bool:
    """Return True if a row already exists in Shopify, following the requested key rules."""
    v = _norm(vendor)
    s = _norm(sku)
    u = _norm_upc(upc)

    # 1) SKU + UPC
    if s and u and (s, u) in key_sets["sku_upc"]:
        return True

    # 2) UPC only
    if u and (u,) in key_sets["upc"]:
        return True

    # 3) Vendor + SKU
    if v and s and (v, s) in key_sets["vendor_sku"]:
        return True

    return False
def _apply_red_font_for_handle(buffer: io.BytesIO, sheet_name: str, rows_to_color: list[int]) -> io.BytesIO:
    """Color the Handle cell red for the given 0-based row indexes (dataframe rows)."""
    wb = load_workbook(buffer)
    ws = wb[sheet_name]

    # locate Handle column
    headers = [str(c.value or "") for c in ws[1]]
    try:
        handle_col_idx = headers.index("Handle") + 1
    except ValueError:
        # nothing to do
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out

    red_font = Font(color="FF0000")
    for df_i in rows_to_color:
        excel_row = df_i + 2  # header +1, df row offset
        cell = ws.cell(row=excel_row, column=handle_col_idx)
        cell.font = red_font

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out
import io
import re
import math
import unicodedata
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


def _sanitize_nan(df):
    """Replace NaN / None with empty string for Shopify export."""
    return df.where(df.notna(), "")

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

def _norm(x) -> str:
    """Normalize input to clean string; treat NaN/None/'nan'/'none' as empty."""
    if x is None:
        return ""
    # pandas/numpy NaN
    if isinstance(x, float) and math.isnan(x):
        return ""
    s = str(x).replace("\u00A0", " ").strip()
    if s.lower() in ("nan", "none"):
        return ""
    return s

# ---------------------------------------------------------
# ORDRE FINAL DES COLONNES (strict)
# ---------------------------------------------------------
SHOPIFY_OUTPUT_COLUMNS = [
    "Handle",
    "Command",
    "Title",
    "Body (HTML)",
    "Vendor",
    "Custom Product Type",
    "Tags",
    "Published",
    "Published Scope",
    "Option1 Name",
    "Option1 Value",
    "Variant SKU",
    "Variant Barcode",
    "Variant Country of Origin",
    "Variant HS Code",
    "Variant Grams",
    "Variant Inventory Tracker",
    "Variant Inventory Policy",
    "Variant Fulfillment Service",
    "Variant Price",
    "Variant Requires Shipping",
    "Variant Taxable",
    "SEO Title",
    "SEO Description",
    "Variant Weight Unit",
    "Cost per item",
    "Status",
    "Metafield: my_fields.product_use_case [multi_line_text_field]",
    "Metafield: my_fields.product_features [multi_line_text_field]",
    "Metafield: my_fields.behind_the_brand [multi_line_text_field]",
    "Metafield: my_fields.size_comment [single_line_text_field]",
    "Metafield: my_fields.gender [single_line_text_field]",
    "Metafield: my_fields.colour [single_line_text_field]",
    "Metafield: mm-google-shopping.color",
    "Variant Metafield: mm-google-shopping.size",
    "Metafield: mm-google-shopping.size_system",
    "Metafield: mm-google-shopping.condition",
    "Metafield: mm-google-shopping.google_product_category",
    "Metafield: mm-google-shopping.gender",
    "Variant Metafield: mm-google-shopping.mpn",
    "Variant Metafield: mm-google-shopping.gtin",
    "Metafield: theme.siblings [single_line_text_field]",
    "Category: ID",
    "Inventory Available: Boutique",
    "Inventory Available: Le Club",
]

# ---------------------------------------------------------
# Helpers
# ---------------------------------------------------------
def _norm(s) -> str:
    s = str(s or "").strip()
    # Treat numeric zeros coming from supplier sheets as empty
    if s in ("0", "0.0", "0.00"):
        return ""
    return re.sub(r"\s+", " ", s)



def _norm_key(s) -> str:
    """Key normalization: trim, lowercase, remove accents."""
    t = _norm(s).lower()
    t = "".join(c for c in unicodedata.normalize("NFKD", t) if not unicodedata.combining(c))
    return t




def _sanitize_text_like_html(v) -> str:
    """Remove HTML-ish artifacts (<br>, &nbsp;) and normalize whitespace."""
    if v is None:
        return ""
    s = str(v)
    # Normalize common HTML breaks to newlines
    s = re.sub(r"(?i)<br\s*/?>", "\n", s)
    # Convert non-breaking spaces
    s = s.replace("&nbsp;", " ").replace("\xa0", " ")
    # Collapse excessive spaces but keep newlines
    s = re.sub(r"[ \t\f\v]+", " ", s)
    # Clean up multiple newlines
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()

def _strip_made_in(s: str) -> str:
    t = _norm(s)
    # remove common prefixes like "Made In "
    t = re.sub(r"(?i)^made\s+in\s+", "", t).strip()
    return t
def _norm_handle(v) -> str:
    s = str(v or "").strip().lower()
    # collapse whitespace
    s = re.sub(r"\s+", "", s)
    return s


def _remove_size_from_handle(handle: str) -> str:
    """
    Supprime toute grandeur À LA FIN du handle.
    IMPORTANT : agit UNIQUEMENT sur le handle.

    Exemples supprimés :
    - -xs, -s, -m, -l, -xl, -xxl, -xxxl
    - -6, -6.5, -10, -10-5, etc.
    """
    if not handle:
        return ""

    h = str(handle).strip().lower()

    # Tailles alpha à la fin
    h = re.sub(r"-(xs|s|m|l|xl|xxl|xxxl)$", "", h)

    # Tailles numériques à la fin (6, 6.5, 10-5, etc.)
    h = re.sub(r"-\d+([.-]\d+)?$", "", h)

    return h


def _strip_gender_prefix_size(v: str) -> str:
    s = _norm(v)
    if not s:
        return ""
    if re.match(r"^[WwMm]\s*\d", s):
        return s[1:].strip()
    return s


def _is_onesize(v: str) -> bool:
    """Return True if the size represents a One Size / OS variant."""
    s = _norm(v).lower()
    if not s:
        return False
    s2 = re.sub(r"[\s\-_]+", "", s)
    if s2 in {"os", "onesize"}:
        return True
    if s.startswith("one size"):
        return True
    if s in {"one-size", "one size", "one_size", "onesize"}:
        return True
    if s in {"o/s", "o-s"}:
        return True
    if "one size fits" in s:
        return True
    return False


def _strip_gender_tokens(text: str) -> str:
    """Remove embedded gender markers like -w-, - W -, -m-, etc from a string."""
    s = str(text or "")
    # remove patterns like -w- , - W - , /w/ etc surrounded by dashes/spaces
    s = re.sub(r"(?i)(\s*-\s*[wm]\s*-\s*)", " ", s)
    s = re.sub(r"(?i)(\b[wm]\b)", lambda m: "" if m.group(0).lower() in ("w","m") else m.group(0), s)
    s = re.sub(r"\s+", " ", s).strip()
    return s




def _clean_style_key(v) -> str:
    s = _norm(v)
    # if Excel treated numeric as float: 123.0 -> 123
    s = re.sub(r"^(\d+)\.0+$", r"\1", s)
    return s

def _strip_reg_for_handle(s: str) -> str:
    """Handle only: remove ® and (r)/[r] to keep URL safe."""
    t = _norm(s)
    t = t.replace("®", "")
    t = re.sub(r"[\(\[\{]\s*r\s*[\)\]\}]", "", t, flags=re.IGNORECASE)
    return _norm(t)


def _convert_r_to_registered(s: str) -> str:
    """Display/SEO: convert (r)/[r] to ®."""
    t = _norm(s)
    t = re.sub(r"[\(\[\{]\s*r\s*[\)\]\}]", "®", t, flags=re.IGNORECASE)
    return t


def _title_case_preserve_registered(text: str) -> str:
    """
    Strict Title Case while preserving ® and ™ (and standalone TM).
    - Title-cases each space-separated token
    - Also title-cases sub-tokens split by "/" and "-"
    - Keeps tokens containing digits as-is
    - Keeps tokens that are exactly TM as "TM"
    """
    text = _norm(text)
    if not text:
        return ""

    def _tc_token(tok: str) -> str:
        if not tok:
            return tok

        # Preserve standalone TM token
        if tok.strip().lower() == "tm":
            return "TM"

        # Keep tokens containing digits as-is
        if any(ch.isdigit() for ch in tok):
            return tok

        # Preserve ® and ™ inside token
        for sym in ("®", "™"):
            if sym in tok:
                sub = tok.split(sym)
                sub = [(_tc_token(s) if s else "") for s in sub]
                return sym.join(sub)

        # separators inside token
        for sep in ["/", "-"]:
            if sep in tok:
                parts = tok.split(sep)
                parts = [(_tc_token(p) if p else "") for p in parts]
                return sep.join(parts)

        return tok[:1].upper() + tok[1:].lower()

    return " ".join(_tc_token(w) for w in text.split(" "))


def _normalize_match_text(s: str) -> str:
    """
    Normalization used ONLY for matching category/product type:
    - tee -> tshirt
    - t-shirt / t shirt -> tshirt
    - long-sleeve / long sleeve -> long sleeve
    """
    t = str(s or "")
    t = t.replace("®", "")
    t = t.lower()

    t = re.sub(r"\bt\s*[- ]\s*shirt\b", "tshirt", t)
    t = re.sub(r"\btshirt\b", "tshirt", t)

    t = re.sub(r"\btee\b", "tshirt", t)
    t = re.sub(r"\btees\b", "tshirt", t)

    t = re.sub(r"\blong\s*[- ]\s*sleeve\b", "long sleeve", t)
    return t


def _words(s: str) -> list[str]:
    return re.findall(r"[a-z0-9]+", _normalize_match_text(s))


def _singularize_token(tok: str) -> str:
    if tok.endswith("s") and len(tok) >= 4:
        return tok[:-1]
    return tok


def _wordset_loose(s: str) -> set[str]:
    return set(_singularize_token(t) for t in _words(s))


def _first_existing_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    cols = {c.lower(): c for c in df.columns}
    for c in candidates:
        if c.lower() in cols:
            return cols[c.lower()]
    return None

def _first_existing_col_with_data(df: pd.DataFrame, candidates: list[str]) -> str | None:
    """Return first candidate column that exists AND has at least one non-empty value."""
    cols = {c.lower(): c for c in df.columns}
    for cand in candidates:
        key = cand.lower()
        if key in cols:
            col = cols[key]
            s = df[col]
            # normalize empties without turning NaN into 'nan'
            s_clean = s.fillna("").astype(str).str.strip()
            # treat literal tokens 'nan'/'none' as empty too
            s_clean = s_clean.replace({r"(?i)^\s*(nan|none)\s*$": ""}, regex=True)
            if s_clean.ne("").any():
                return col
    return None



# ---------------------------------------------------------
# Help data readers (openpyxl)
# ---------------------------------------------------------
def _load_help_wb(help_bytes: bytes):
    return openpyxl.load_workbook(io.BytesIO(help_bytes), data_only=True)


def _read_2col_map(wb, sheet_candidates: list[str]) -> dict[str, str]:
    """Col A raw -> Col B standard"""
    sheet = None
    for name in sheet_candidates:
        if name in wb.sheetnames:
            sheet = wb[name]
            break
    if sheet is None:
        return {}

    m: dict[str, str] = {}
    for r in range(2, sheet.max_row + 1):
        a = sheet.cell(row=r, column=1).value
        b = sheet.cell(row=r, column=2).value
        if a is None or b is None:
            continue
        ra = str(a).strip()
        rb = str(b).strip()
        if not ra or ra.lower() == "nan":
            continue
        m[ra.lower()] = rb
    return m


def _standardize(val: str, mapping: dict[str, str]) -> str:
    s = _norm(val)
    if not s or s.lower() == "nan":
        return ""
    return mapping.get(s.lower(), s)

def _build_country_code_map(mapping: dict[str, str]) -> dict[str, str]:
    """
    Build a more permissive country -> ISO-2 code map from Help Data.
    Help data often contains official names like 'Moldova (the Republic of)'.
    We add simplified keys (remove parentheses content, commas, and trailing '(the ...)' patterns)
    so inputs like 'Moldova' still resolve to 'MD'.
    """
    out: dict[str, str] = {}
    for k, v in (mapping or {}).items():
        if not k:
            continue
        key = str(k).strip().lower()
        code = str(v or "").strip()
        if not code:
            continue
        out[key] = code

        # simplified version: remove parenthetical content, commas, and extra spaces
        simp = re.sub(r"\s*\(.*?\)\s*", " ", key)
        simp = simp.replace(",", " ")
        simp = re.sub(r"\s+", " ", simp).strip()
        if simp:
            out.setdefault(simp, code)

        # also drop trailing ' (the ...' already handled by parentheses removal; and 'the' articles
        simp2 = re.sub(r"\bthe\b", "", simp).strip()
        simp2 = re.sub(r"\s+", " ", simp2).strip()
        if simp2:
            out.setdefault(simp2, code)

    return out


def _standardize_country(val: str, country_code_map: dict[str, str]) -> str:
    """
    Standardize Country of Origin to ISO-2 codes expected by Shopify.
    - Accepts already-coded values (2 letters)
    - Uses permissive matching against help data (simplified keys)
    """
    s = _norm(val)
    if not s or s.lower() == "nan":
        return ""
    t = s.strip()
    # If already looks like ISO-2 code
    if re.fullmatch(r"[A-Za-z]{2}", t):
        return t.upper()

    key = t.lower()
    if key in country_code_map:
        return country_code_map[key]

    simp = re.sub(r"\s*\(.*?\)\s*", " ", key)
    simp = simp.replace(",", " ")
    simp = re.sub(r"\s+", " ", simp).strip()
    if simp in country_code_map:
        return country_code_map[simp]

    simp2 = re.sub(r"\bthe\b", "", simp).strip()
    simp2 = re.sub(r"\s+", " ", simp2).strip()
    return country_code_map.get(simp2, s)



def _read_list_column(wb, sheet_name: str) -> list[str]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    out = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is None:
            continue
        s = str(v).strip()
        if s and s.lower() != "nan":
            out.append(s)
    return out



def _read_product_type_gendered_map(wb, sheet_name: str = "Product Types") -> dict[str, bool]:
    """Read Product Types sheet to know if a Custom Product Type is gendered.

    Expected sheet structure (as in Help Data):
      Col A: Custom Product Type
      Col B: 'Genré' or 'NON Genré' (can be blank)

    Returns a dict[normalized_product_type -> is_gendered_bool].

    Normalization:
      - lowercased + trimmed
      - we also add simple singular/plural variants so that:
          "Water Bottle" (help data) matches "Water Bottles" (output)
          "Vest" matches "Vests", etc.

    Rules:
      - If Col B contains 'non' => NON Genré => False
      - Otherwise => True (keep previous behavior)
    """
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]

    def _norm_pt(x: str) -> str:
        s = str(x or "").strip().lower()
        s = re.sub(r"\s+", " ", s).strip()
        return s

    def _singularize(s: str) -> str:
        # very small heuristic: Water Bottles -> Water Bottle, Vests -> Vest, etc.
        t = _norm_pt(s)
        if t.endswith("s") and len(t) >= 4 and not t.endswith("ss"):
            return t[:-1]
        return t

    def _pluralize(s: str) -> str:
        t = _norm_pt(s)
        if not t:
            return t
        if t.endswith("s"):
            return t
        return t + "s"

    m: dict[str, bool] = {}
    for r in range(2, ws.max_row + 1):
        pt = ws.cell(row=r, column=1).value
        flag = ws.cell(row=r, column=2).value
        if pt is None:
            continue
        pt_s = str(pt).strip()
        if not pt_s or pt_s.lower() == "nan":
            continue

        flag_s = "" if flag is None else str(flag).strip().lower()
        # normalize accents (minimal)
        flag_s = flag_s.replace("é", "e").replace("è", "e").replace("ê", "e").replace("à", "a").replace("ô", "o")

        is_gendered = False if ("non" in flag_s) else True

        base = _norm_pt(pt_s)
        if not base:
            continue

        # store base + variants
        keys = {base, _singularize(base), _pluralize(base), _pluralize(_singularize(base))}
        for k in keys:
            if k:
                m[k] = is_gendered

    return m



def _read_variant_weight_map(wb, sheet_name: str = "Variant Weight (Grams)") -> dict[str, str]:
    """
    Map Custom Product Type -> Variant Weight (Grams)
    from Help Data sheet "Variant Weight (Grams)".
    """
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    m: dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        k = ws.cell(row=r, column=1).value
        v = ws.cell(row=r, column=2).value
        if k is None or v is None:
            continue
        ks = str(k).strip()
        if not ks or ks.lower() == "nan":
            continue
        # keep grams as string (Shopify expects a number; empty string means "unknown")
        if isinstance(v, (int, float)) and not (isinstance(v, float) and math.isnan(v)):
            vs = str(int(v)) if float(v).is_integer() else str(v)
        else:
            vs = str(v).strip()
        if vs.lower() == "nan":
            continue
        m[ks.lower()] = vs
    return m

def _read_category_rows(wb, sheet_name: str):
    """returns list[(name_keywords, id)] from columns A,B. Handles sheets with or without headers."""
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]

    # Detect header row (light heuristic)
    a1 = ws.cell(row=1, column=1).value
    b1 = ws.cell(row=1, column=2).value
    start_row = 1
    if isinstance(a1, str) and a1.strip().lower() in {"name", "keyword", "category", "product category"}:
        start_row = 2
    if isinstance(b1, str) and b1.strip().lower() in {"id", "category id"}:
        start_row = 2

    rows = []
    for r in range(start_row, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        if a is None:
            continue
        aa = str(a).strip()
        bb = "" if b is None else str(b).strip()
        if aa and aa.lower() != "nan":
            rows.append((aa, bb))
    return rows or None


def _read_brand_line_map(wb, sheet_name: str) -> dict[str, str]:
    """Col A = brand, Col B+ concatenated text parts"""
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    m = {}
    for r in range(2, ws.max_row + 1):
        brand = ws.cell(row=r, column=1).value
        if brand is None:
            continue
        b = str(brand).strip()
        if not b or b.lower() == "nan":
            continue

        parts = []
        for c in range(2, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v).strip()
            if s and s.lower() != "nan":
                parts.append(s)

        if parts:
            # IMPORTANT: keep exact spacing; caller decides punctuation
            m[b.lower()] = " ".join(parts).strip()
    return m


def _read_size_reco_map(wb) -> dict[str, str]:
    """Garment -> Comment"""
    if "Size Recommandation" not in wb.sheetnames:
        return {}
    ws = wb["Size Recommandation"]

    headers = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if h is None:
            continue
        headers[str(h).strip().lower()] = c

    gcol = headers.get("garment")
    ccol = headers.get("comment")
    if not gcol or not ccol:
        return {}

    m = {}
    for r in range(2, ws.max_row + 1):
        g = ws.cell(row=r, column=gcol).value
        c = ws.cell(row=r, column=ccol).value
        if g is None or c is None:
            continue
        gs = str(g).strip()
        cs = str(c).strip()
        if gs and cs:
            m[gs.lower()] = cs
    return m


# ---------------------------------------------------------
# Matching functions
# ---------------------------------------------------------
def _best_match_id(text: str, cat_rows) -> str:
    """
    Exact-match (loose singular/plural): all words in name must be in text.
    Returns ID (col B).

    Special rule:
    - If text contains "long sleeve" but no specific garment match is found,
      ALWAYS try "long sleeve jersey" (never tshirt).
    """
    if not cat_rows:
        return ""

    def _match(t: str) -> str:
        tset = _wordset_loose(t)
        best_id = ""
        best_len = 0
        for name, cid in cat_rows:
            nset = _wordset_loose(name)
            if nset and nset.issubset(tset):
                if len(nset) > best_len:
                    best_len = len(nset)
                    best_id = str(cid or "").strip()
        best_id = re.sub(r"\.0$", "", best_id) if best_id else ""
        return best_id

    # 1) normal match
    got = _match(text)
    if got:
        return got

    # 2) LONG SLEEVE fallback -> ALWAYS Jersey
    w = _wordset_loose(text)
    if {"long", "sleeve"}.issubset(w):
        got = _match(f"{text} jersey")
        if got:
            return got

    return ""

    def _match(t: str) -> str:
        tset = _wordset_loose(t)
        best_id = ""
        best_len = 0
        for name, cid in cat_rows:
            nset = _wordset_loose(name)
            if nset and nset.issubset(tset):
                if len(nset) > best_len:
                    best_len = len(nset)
                    best_id = str(cid or "").strip()
        best_id = re.sub(r"\.0$", "", best_id) if best_id else ""
        return best_id

    # 1) normal match
    got = _match(text)
    if got:
        return got

    # 2) LONG SLEEVE fallback: if only "long sleeve" appears, we try common garment types
    w = _wordset_loose(text)
    if {"long", "sleeve"}.issubset(w):
        # Prefer T-Shirt when no further hint is present
        got = _match(f"{text} tshirt")
        if got:
            return got
        got = _match(f"{text} jersey")
        if got:
            return got

    return ""


def _best_match_product_type(text: str, product_types: list[str]) -> str:
    """
    Match product type by word-subset (loose singular/plural).

    Special rule:
    - If text contains "long sleeve" but no specific garment match is found,
      ALWAYS try "long sleeve jersey" (never tshirt).
    """
    def _match(t: str) -> str:
        tset = _wordset_loose(t)
        best = ""
        best_len = 0
        for pt in product_types:
            pset = _wordset_loose(pt)
            if pset and pset.issubset(tset):
                if len(pset) > best_len:
                    best_len = len(pset)
                    best = pt
        return best

    # 1) normal match
    got = _match(text)
    if got:
        return got

    # 2) LONG SLEEVE fallback -> ALWAYS Jersey
    w = _wordset_loose(text)
    if {"long", "sleeve"}.issubset(w):
        got = _match(f"{text} jersey")
        if got:
            return got

    return ""


# ---------------------------------------------------------
# Parsing & formatting
# ---------------------------------------------------------
def _extract_color_size_from_description(desc: str) -> tuple[str, str]:
    text = _norm(desc)
    if not text:
        return "", ""
    parts = re.split(r"\s*[-,/]\s*|\s*,\s*", text)
    parts = [p.strip() for p in parts if p and p.strip()]
    if len(parts) < 2:
        return "", ""
    last = parts[-1]
    if re.fullmatch(r"(X{0,3}S|X{0,3}L|S|M|L|XL|XXL|XXXL|\d{1,2}([./-]\d{1,2})?)", last, flags=re.IGNORECASE):
        return parts[-2], last
    return parts[-1], ""


def _round_to_nearest_9_99(price) -> float:
    """
    Pricing rule (ALL CAD suppliers):
    - Round to the nearest dollar (half-up)
    - Then subtract 0.01 (ex: 115 -> 114.99)
    - If result would be <= 0, return NaN (caller will blank the cell)
    """
    if price is None or (isinstance(price, float) and math.isnan(price)):
        return float("nan")
    try:
        p = float(price)
    except Exception:
        return float("nan")

    # Nearest dollar, half-up
    dollar = math.floor(p + 0.5)
    val = round(dollar - 0.01, 2)

    if val <= 0:
        return float("nan")
    return val
def _barcode_keep_zeros(x) -> str:
    """Normalize barcode/UPC/EAN.
    - Keep digits only when the value is numeric.
    - Preserve leading zeros for UPC (pad to 12 when length <= 12).
    - Accept EAN/other barcodes up to 16 digits (kept as-is, no padding).
    """
    if x is None:
        return ""
    s = str(x).strip()
    if s == "" or s.lower() == "nan" or s in ("0", "0.0", "0.00"):
        return ""
    # Excel floats like 123.0
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]

    # Keep digits only if it's mostly numeric
    digits = re.sub(r"\D", "", s)
    if digits == "":
        return s

    # Treat 0-only barcodes as empty
    try:
        if int(digits) == 0:
            return ""
    except Exception:
        pass

    if len(digits) <= 12:
        return digits.zfill(12)
    if len(digits) <= 16:
        return digits
    return digits[:16]


def _hs_code_clean(x) -> str:
    """Clean HS/HTS code and keep only the first 6 characters (no dots)."""
    if x is None:
        return ""
    s = str(x).strip()
    if s == "" or s.lower() == "nan" or s in ("0", "0.0", "0.00"):
        return ""
    s = re.sub(r"\.0$", "", s)
    # Remove separators (dots/spaces/etc.)
    s = re.sub(r"[^0-9A-Za-z]", "", s)
    return s[:6]


# ---------------------------------------------------------
# Excel highlighting
# ---------------------------------------------------------
def _apply_yellow_for_empty(buffer: io.BytesIO, sheet_name: str, cols_to_yellow: list[str]) -> io.BytesIO:
    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb[sheet_name]

    header = [cell.value for cell in ws[1]]
    col_index = {name: i + 1 for i, name in enumerate(header) if name}

    for col_name in cols_to_yellow:
        if col_name not in col_index:
            continue
        c = col_index[col_name]
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            if val is None or (isinstance(val, str) and val.strip() == ""):
                cell.fill = YELLOW_FILL

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

def _apply_red_font_for_rows_cols(buffer: io.BytesIO, sheet_name: str, rows_0based: list[int], col_names: list[str]) -> io.BytesIO:
    """Apply red font to specific columns for the given 0-based dataframe row indexes."""
    buffer.seek(0)
    wb = load_workbook(buffer)
    if sheet_name not in wb.sheetnames:
        return buffer
    ws = wb[sheet_name]

    headers = [str(c.value or "") for c in ws[1]]
    col_index = {h: i + 1 for i, h in enumerate(headers) if h}

    red_font = Font(color="FF0000")
    for df_i in rows_0based:
        excel_row = df_i + 2
        for cn in col_names:
            if cn not in col_index:
                continue
            cell = ws.cell(row=excel_row, column=col_index[cn])
            cell.font = red_font

    outb = io.BytesIO()
    wb.save(outb)
    outb.seek(0)
    return outb


def _apply_header_notes(buffer: io.BytesIO, sheet_name: str, notes: dict[str, str]) -> io.BytesIO:
    """
    Add an Excel comment (note) on the HEADER CELL for specified columns.
    Applied to the sheet's row 1 only.
    """
    buffer.seek(0)
    wb = load_workbook(buffer)
    if sheet_name not in wb.sheetnames:
        return buffer
    ws = wb[sheet_name]

    headers = [str(c.value or "") for c in ws[1]]
    col_index = {h: i + 1 for i, h in enumerate(headers) if h}

    for col_name, note in notes.items():
        if col_name not in col_index:
            continue
        cell = ws.cell(row=1, column=col_index[col_name])
        cell.comment = Comment(note, "Le Club")
        cell.comment.width = 360
        cell.comment.height = 120

    outb = io.BytesIO()
    wb.save(outb)
    outb.seek(0)
    return outb

# ---------------------------------------------------------
# MAIN
# ---------------------------------------------------------
def run_transform(
    supplier_xlsx_bytes: bytes,
    help_xlsx_bytes: bytes,
    vendor_name: str,
    brand_choice: str = "",
    event_promo_tag: str = "",
    style_season_map: dict[str, str] | None = None,
    existing_shopify_xlsx_bytes: bytes | None = None,
    supplier_filename: str = "",
):
    # Defensive defaults (avoid NameError when price columns absent)
    detected_cost_col = None
    detected_price_col = None
    warnings: list[dict] = []

    style_season_map = style_season_map or {}
    vendor_key = _colkey(vendor_name)
    is_satisfy = vendor_key in ("satisfy",)
    style_season_map = { _clean_style_key(k): v for k, v in style_season_map.items() }

    # -----------------------------------------------------
    # Supplier reader (multi-sheet capable)
    # -----------------------------------------------------
    def _read_supplier_multi_sheet(file_bytes: bytes, file_name: str = "") -> pd.DataFrame:
        """
        Reads supplier XLSX.
        - If there are multiple sheets, keep only sheets that contain the minimum required columns
          (Description-like), then concatenate.
        - If there is a single valid sheet, behaves like the previous implementation.
        """
        # CSV support (v15): allow suppliers to provide a single CSV instead of XLSX
        if str(file_name or "").strip().lower().endswith(".csv"):
            try:
                df_csv = _read_supplier_csv(io.BytesIO(file_bytes), file_name)
            except Exception as e:
                raise ValueError(f"Impossible de lire le CSV fournisseur: {e}")
            return df_csv

        bio = io.BytesIO(file_bytes)
        xls = pd.ExcelFile(bio)

        # Supplier-specific override: PAS Normal Studios uses only "Summary + Data"
        vendor_key = _colkey(vendor_name)
        if vendor_key in ("pasnormalstudios", "pasnormalstudio"):
            target_sheet = "Summary + Data"
            if target_sheet in xls.sheet_names:
                df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=target_sheet, dtype=str)
                df = df.dropna(how="all")
                if df is None or df.empty:
                    raise ValueError('Onglet "Summary + Data" vide dans le fichier fournisseur.')
                df["_source_sheet"] = target_sheet
                return df


        # Column candidates duplicated from the main logic (kept local to avoid refactors).
        desc_candidates = [
            "description", "Description", "Product Name", "product name",
            "Name", "name",
            "Title", "title", "Style", "style", "Style Name", "style name",
            "Display Name", "display name", "Online Display Name", "online display name",
            "Technical Specifications", "technical specifications",
        ]
        msrp_candidates = [
            "Cad MSRP", "MSRP", "Retail Price (CAD)", "retail price (CAD)", "retail price (cad)",
        ]

        dfs: list[pd.DataFrame] = []
        for sn in xls.sheet_names:
            df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sn, dtype=str)
            # --- Price columns (robust) ---
            cost_col = _find_col(df.columns, [
                "Wholesale CAD", "Wholesale (CAD)", "CAD Wholesale", "WholesaleCAD", "wholesale cad"
            ])
            price_col = _find_col(df.columns, [
                "Retail CAD", "Retail (CAD)", "CAD Retail", "RetailCAD", "retail cad"
            ])

            # Legacy MSRP-like columns (optional)
            msrp_col = _find_col(df.columns, [
                "Retail Price (CAD)", "Cad MSRP", "MSRP", "msrp"
            ])

            # Prefer explicit Norda CAD columns
            detected_cost_col = cost_col if cost_col else None
            if price_col:
                detected_price_col = price_col
            elif msrp_col:
                detected_price_col = msrp_col
            else:
                detected_price_col = None

            # Drop fully empty rows early
            if df is None or df.empty:
                warnings.append({
                    "type": "sheet_skipped",
                    "sheet": sn,
                    "reason": "empty",
                })
                continue
            df = df.dropna(how="all")
            if df.empty:
                warnings.append({
                    "type": "sheet_skipped",
                    "sheet": sn,
                    "reason": "empty",
                })
                continue

            # Validate minimum required columns
            has_desc = _first_existing_col(df, desc_candidates) is not None
            has_msrp = _first_existing_col(df, msrp_candidates) is not None
            if not has_desc:
                warnings.append({
                    "type": "sheet_skipped",
                    "sheet": sn,
                    "reason": "missing_required_columns",
                    "has_desc": has_desc,
                    "has_msrp": has_msrp,
                })
                continue

            df["_source_sheet"] = sn
            dfs.append(df)

        if not dfs:
            raise ValueError(
                "Aucun onglet valide détecté dans le fichier fournisseur (colonne Description requise)."
            )
        return pd.concat(dfs, ignore_index=True, sort=False)

    sup = _read_supplier_multi_sheet(supplier_xlsx_bytes, supplier_filename).copy()

    # Satisfy: remove Totals line (often contains zeros that should not become products)
    if vendor_key in ("satisfy",):
        name_col_tmp = _first_existing_col(sup, ["name", "Name"])
        if name_col_tmp:
            sup = sup.loc[~sup[name_col_tmp].astype(str).str.strip().str.lower().eq("totals")].copy()

    # PAS Normal Studios: keep only rows with Order Qty >= 1 (from "Summary + Data")
    if vendor_key in ("pasnormalstudios", "pasnormalstudio"):
        order_qty_col = _first_existing_col(sup, ["Order Qty", "order qty", "Order Quantity", "order quantity"])
        if order_qty_col:
            qty_num = pd.to_numeric(sup[order_qty_col].astype(str).str.replace(",", "", regex=False).str.strip(), errors="coerce").fillna(0)
            before_n = len(sup)
            sup = sup.loc[qty_num >= 1].copy()
            after_n = len(sup)
            if after_n < before_n:
                warnings.append({"type": "rows_filtered", "reason": "Order Qty < 1", "removed": before_n - after_n})
        else:
            warnings.append({"type": "missing_column", "column": "Order Qty", "vendor": vendor_name})



    # -----------------------------------------------------
    # Detect price columns on the concatenated supplier dataframe
    # (fix: detected_* set inside sheet loop are not propagated)
    # -----------------------------------------------------
    detected_cost_col = _find_col(sup.columns, [
        "Wholesale CAD", "Wholesale (CAD)", "CAD Wholesale", "WholesaleCAD", "wholesale cad",
        "Landed", "landed", "Wholesale Price (CAD)", "wholesale price (cad)", "Wholesale Price", "wholesale price"
    ])
    detected_price_col = _find_col(sup.columns, [
        "Retail CAD", "Retail (CAD)", "CAD Retail", "RetailCAD", "retail cad",
        "Retail Price (CAD)", "Cad MSRP", "MSRP", "msrp"
    ])

    
    # Satisfy: the file uses generic columns (retail/retail price/price and wholesale) without "CAD" in header.
    # We override detection so Variant Price / Cost per item are populated for this supplier only.
    if is_satisfy:
        detected_price_col = _find_col(sup.columns, ["retail price", "retail", "price"])
        detected_cost_col = _find_col(sup.columns, ["wholesale"])
    wb = _load_help_wb(help_xlsx_bytes)

    # Standardization
    color_map = _read_2col_map(wb, ["Color Standardization", "Color Variable"])
    size_map = _read_2col_map(wb, ["Size Standardization", "Size Variante"])
    country_map = _read_2col_map(wb, ["Country Abbreviations", "Country of Origin"])
    country_code_map = _build_country_code_map(country_map)
    gender_map = _read_2col_map(wb, ["Gender Standardization", "Gender"])

    # Categories & Product types
    shopify_cat_rows = _read_category_rows(wb, "Shopify Product Category")
    google_cat_rows = _read_category_rows(wb, "Google Product Category")
    product_types = _read_list_column(wb, "Product Types")

    # Canonical Product Type resolver:
    # Output Custom Product Type MUST match EXACTLY one value from Help Data -> Product Types (col A).
    _pt_canon = { _norm_key(pt): pt for pt in (product_types or []) if _norm(pt) }

    def _canon_product_type(pt: str) -> str:
        s = _norm(pt)
        if not s:
            return ""
        k = _norm_key(s)
        if k in _pt_canon:
            return _pt_canon[k]
        # singular/plural tolerance (only for matching; returns canonical from Help Data)
        if k.endswith("s") and k[:-1] in _pt_canon:
            return _pt_canon[k[:-1]]
        if (k + "s") in _pt_canon:
            return _pt_canon[k + "s"]
        return s

    product_type_gendered_map = _read_product_type_gendered_map(wb, "Product Types")
    variant_weight_map = _read_variant_weight_map(wb)


    # Brand maps
    brand_desc_map = _read_brand_line_map(wb, "SEO Description Brand Part")
    brand_lines_map = _read_brand_line_map(wb, "Brand lines")

    # Size reco
    size_comment_map = _read_size_reco_map(wb)

    # Supplier columns
    desc_col = _first_existing_col_with_data(
        sup,
        [
            "Description", "description",
            "Style Description", "style description",
            "Product Details", "product details",
            "Technical Specifications", "technical specifications",
            "Product Name", "product name",
            "Title", "title", "Style", "style", "Style Name", "style name",
            "Name", "name",
            "Display Name", "display name", "Online Display Name", "online display name",
        ],
    )

    # If we picked Technical Specifications but it is mostly empty, fallback to Description when available.
    desc_col_fallback = _first_existing_col(sup, ["Description", "description"])
    if desc_col and _colkey(desc_col) in ("technicalspecifications", "technicalspecification") and desc_col_fallback:
        non_empty_ratio = _series_str_clean(sup[desc_col]).str.strip().ne("").mean() if len(sup) else 0
        if non_empty_ratio < 0.2:
            desc_col = desc_col_fallback

    product_col = _first_existing_col(sup, ["Product", "Product Code", "SKU", "sku"])
    color_col = _first_existing_col(sup, ["Vendor Color", "vendor color", "Color", "color", "Colour", "colour", "Color Code", "color code", "colour code and name", "Colour Code and Name", "Color Code and Name"])
    size_col = _first_existing_col(sup, ["Size 1","Size1","Size", "size", "Vendor Size1", "vendor size1"])
    upc_col = _first_existing_col(sup, ["UPC", "UPC Code", "UPC Code.", "UPC Code 1", "UPC Code1", "UPC1", "Variant Barcode", "Barcode", "bar code", "upc", "upc code"])
    ean_col = _first_existing_col(sup, ["EAN", "EAN Code", "ean", "ean code"])
    origin_col = _first_existing_col(sup, ["Country of origin", "Country of Origin", "Country Of Origin", "Country Code", "Origin", "Manufacturing Country", "COO", "country of origin", "country of origin ", "country code", "origin", "manufacturing country", "coo"])
    hs_col = _first_existing_col(sup, ["HS Code", "HTS Code", "hs code", "hts code", "commodity hs", "commodity hts", "Commodity HS", "Commodity HTS", "custome tarif code (no dots)", "custom tarif code (no dots)", "custom tarif code", "Custom tarif code (no dots)", "Custom tarif code", "custom tariff code (no dots)", "custom tariff code", "tariff code"])
    extid_col = _first_existing_col(sup, ["External ID", "ExternalID"])
    msrp_col = _first_existing_col(sup, ["Cad MSRP", "MSRP", "Retail Price (CAD)", "retail price (CAD)", "retail price (cad)"])
    landed_col = _first_existing_col(sup, ["Landed", "landed", "Wholesale Price", "wholesale price", "Wholesale Price (CAD)", "wholesale price (cad)"])
    grams_col = _first_existing_col(sup, ["Grams", "Weight (g)", "Weight"])
    gender_col = _first_existing_col(sup, ["Gender", "gender", "Genre", "genre", "Sex", "sex", "Sexe", "sexe"])


    # -----------------------------------------------------
    # Gender inference: detect "-w-" / "- W -" / "-m-" / "- M -" in Name or SKU
    # -----------------------------------------------------
    name_hint_col = _first_existing_col(sup, ["Style Name", "Name", "Product Name", "Title", "Style", "Description", "Display Name", "Online Display Name"])
    sku_hint_col = extid_col or product_col
    def _infer_gender_from_texts(name_val: str, sku_val: str) -> str:
        # Look across name/description-like text + sku for gender signals
        t = f"{_norm(name_val)} {_norm(sku_val)}".lower()

        # Strong markers in SKUs like -w- / -m-
        if re.search(r"-\s*w\s*-", t):
            return "Women"
        if re.search(r"-\s*m\s*-", t):
            return "Men"

        # Text markers (women/men, women's/men's)
        if re.search(r"\bwomen\b|\bwomen's\b|\bwomens\b|\bfemale\b", t):
            return "Women"
        if re.search(r"\bmen\b|\bmen's\b|\bmens\b|\bmale\b", t):
            return "Men"
        return ""

    if desc_col is None:
        raise ValueError(
            "Colonne Description introuvable. Colonnes acceptées: Description, Style, Style Name, Name, Product Name, Title, Display Name, Online Display Name."
        )
    if msrp_col is None:
        msrp_col = None  # MSRP not found; leave prices blank per rules

    # -----------------------------------------------------
    # De-duplicate across sheets (SKU and/or UPC)
    # -----------------------------------------------------
    def _clean_sku(x) -> str:
        s = _norm(x)
        return re.sub(r"\.0$", "", s)

    def _make_dedupe_key(r) -> str:
        sku = _clean_sku(r.get(extid_col, "")) if extid_col else ""
        if not sku:
            sku = _clean_sku(r.get(product_col, "")) if product_col else ""
        upc = _barcode_keep_zeros(r.get(upc_col, "")) if upc_col else ""

        if sku and upc:
            return f"{sku}|{upc}"
        if sku:
            return sku
        if upc:
            return upc
        # Fallback (rare): keep row unique if neither exists
        return ""

    sup["_dedupe_key"] = sup.apply(_make_dedupe_key, axis=1)

    before = len(sup)
    # Only dedupe rows where we have at least one identifier; keep all others.
    has_key = sup["_dedupe_key"].astype(str).str.strip().ne("")
    sup_keyed = sup.loc[has_key].drop_duplicates(subset=["_dedupe_key"], keep="first")
    sup_unkeyed = sup.loc[~has_key]
    sup = pd.concat([sup_keyed, sup_unkeyed], ignore_index=True)
    after = len(sup)
    if after < before:
        warnings.append({
            "type": "dedupe_applied",
            "dedupe_by": "SKU and/or UPC",
            "rows_before": before,
            "rows_after": after,
            "rows_removed": before - after,
        })

    
    # Base description (keep both a normalized version and the original source text)
    sup["_desc_source"] = _series_str_clean(sup[desc_col])  # preserve original (length, punctuation, line breaks)
    sup["_desc_raw"] = sup["_desc_source"].map(_norm)
    sup["_desc_seo"] = sup["_desc_raw"].apply(_convert_r_to_registered)
    sup["_desc_handle"] = sup.apply(lambda r: _strip_reg_for_handle(r["_title_name_raw"]) if r.get("_desc_is_long") and r.get("_title_name_raw") else _strip_reg_for_handle(r["_desc_raw"]), axis=1)

    # -----------------------------------------------------
    # Long description rule:
    # If the SOURCE description text is > 200 chars, move it to Body (HTML)
    # and build Title from Style Name / Name instead of the long description.
    # -----------------------------------------------------
    title_name_col = _first_existing_col(sup, ["Style Name", "Name", "Product Name", "Title", "Style"])
    sup["_title_name_raw"] = _series_str_clean(sup[title_name_col]).map(_norm) if title_name_col else ""

    sup["_desc_is_long"] = sup["_desc_source"].apply(lambda x: len(str(x)) > 200)

    # Put the original description in Body (HTML) when long (not the normalized one)
    sup["_body_html"] = sup.apply(lambda r: str(r["_desc_source"]).strip() if r["_desc_is_long"] else "", axis=1)
    # Clean HTML-ish artifacts in Body (HTML)
    sup["_body_html"] = sup["_body_html"].map(_sanitize_text_like_html)
    # Color / Size input
    sup["_color_raw"] = _series_str_clean(sup[color_col]).map(_norm) if color_col else ""
    sup["_size_raw"] = _series_str_clean(sup[size_col]).map(_norm) if size_col else ""

    # Fallback parse from description if missing
    parsed = sup["_desc_raw"].apply(_extract_color_size_from_description)
    sup["_color_fb"] = parsed.map(lambda t: t[0])
    sup["_size_fb"] = parsed.map(lambda t: t[1])

    sup["_color_in"] = sup["_color_raw"]

    # PAS Normal Studios – OS / One Size is a size, never a color
    if vendor_key in ("pasnormalstudios", "pasnormalstudio"):
        sup.loc[sup["_color_in"].str.upper().isin(["OS", "ONE SIZE"]), "_color_in"] = ""

    sup.loc[sup["_color_in"].eq(""), "_color_in"] = sup["_color_fb"]
    # PAS Normal Studios – OS / One Size is a size, never a color (applied after fallbacks)
    if vendor_key in ("pasnormalstudios", "pasnormalstudio"):
        sup.loc[sup["_color_in"].astype(str).str.strip().str.upper().isin(["OS", "ONE SIZE"]), "_color_in"] = ""


    sup["_size_in"] = sup["_size_raw"]
    sup.loc[sup["_size_in"].eq(""), "_size_in"] = sup["_size_fb"]

    # Standardize
    sup["_color_std"] = sup["_color_in"].apply(lambda x: _standardize(x, color_map))
    sup["_color_map_hit"] = sup["_color_in"].apply(lambda x: (str(_norm(x)).lower() in set(color_map.keys())) if color_map else True)

    sup["_size_std"] = sup["_size_in"].apply(lambda x: _standardize(x, size_map))

    # Gender (standardize if possible)
    sup["_gender_raw"] = _series_str_clean(sup[gender_col]).map(_norm) if gender_col else ""

    sup["_gender_inferred"] = sup.apply(
        lambda r: _infer_gender_from_texts(
            r.get(name_hint_col, "") if name_hint_col else "",
            r.get(sku_hint_col, "") if sku_hint_col else "",
        ),
        axis=1,
    )
    sup.loc[sup["_gender_raw"].astype(str).str.strip().eq(""), "_gender_raw"] = sup.loc[
        sup["_gender_raw"].astype(str).str.strip().eq(""),
        "_gender_inferred",
    ]

    sup["_gender_std"] = sup["_gender_raw"].apply(lambda x: _standardize(x, gender_map)) if gender_map else sup["_gender_raw"]

    # Vendor / Brand
    sup["_vendor"] = vendor_name
    sup["_brand_choice"] = _norm(brand_choice)

    title_desc_col = None  # init to avoid UnboundLocalError
    # Title: Gender('s) + Description - Color (NON-standardized, Title Cased)
    # -----------------------------------------------------
    # Title rules (kept stable across suppliers)
    # a) Gender ('s if Men/Women) + Description + " - " + Color
    # b) Color NON-standardized (Vendor Color / Color / Colour / Color Code)
    # c) Description from: Description, Product Name, Title, Style, Style Name, Display Name, Online Display Name
    # d) Title Case, ® conserved
    # e) Truncate to max 200 chars
    # -----------------------------------------------------
    
    def _gender_for_title(g: str) -> str:
        """Title prefix rule:
        - ONLY prefix Women's (ensure it appears for women's products)
        - NEVER prefix Men, Unisex, or anything else in Title
        """
        gg = _norm(g)
        if not gg:
            return ""
        ggl = gg.lower().replace("’", "'")
        # Accept common normalized forms (incl. already possessive)
        if ggl in ("women", "womens", "women's", "female", "femme", "femmes"):
            return "Women's"
        return ""

    # Column to use as the primary "name/description" source for Title.
    # IMPORTANT: even when a column is selected, we still do row-level fallbacks
    # (ex: MAAP has a Description column but many rows are empty -> fallback to Name per row).
    title_desc_col = _first_existing_col_with_data(
        sup,
        [
            "Description",
            "Product Name",
            "Title",
            "Style",
            "Style Name",
            "Display Name",
            "Online Display Name",
            "Name",
            "name",
        ],
    )

    # Safeguard: if selected title column yields all-empty, fallback to Name/name.
    if title_desc_col is not None:
        _tmp = _series_str_clean(sup[title_desc_col]).str.strip()
        if (_tmp.eq("").all()) and ("Name" in sup.columns or "name" in sup.columns):
            title_desc_col = _first_existing_col_with_data(sup, ["Name", "name"])

    # SATISFY: prefer supplier Name/name column for Title (same naming basis as handle/SEO title).
    if vendor_key in ("satisfy",):
        _s_name = _first_existing_col(sup, ["Name", "name"])
        if _s_name:
            title_desc_col = _s_name

    # Build description text used for Title (normalized, row-wise fallbacks).
    if title_desc_col is not None:
        _desc_series = _series_str_clean(sup[title_desc_col]).map(_norm)
    else:
        _desc_series = _series_str_clean(sup["_desc_seo"]).map(_norm)

    _name_series = _series_str_clean(sup.get("_title_name_raw", "")).map(_norm)

    # Row-level fallback:
    # - If original supplier description is long (>200 chars) we use Style Name/Name (already in _title_name_raw).
    # - Else if the selected description cell is empty, fallback to Style Name/Name.
    _desc_series = _desc_series.where(_desc_series.astype(str).str.strip().ne(""), _name_series)
    if "_desc_is_long" in sup.columns:
        mask_long = sup["_desc_is_long"] & _name_series.astype(str).str.strip().ne("")
        _desc_series = _desc_series.where(~mask_long, _name_series)

    sup["_desc_title_norm"] = _desc_series.apply(_convert_r_to_registered)
# Clean description text used for Title/SEO fields:
    # - remove embedded gender markers like -w- / -m-
    # - remove leading Men/Women tokens to avoid duplicates with Gender prefix
    def _clean_desc_for_display(s: str) -> str:
        t = _norm(s)
        if not t:
            return ""
        t = _strip_gender_tokens(t)
        # remove leading gender words (men/women/men's/women's)
        t = re.sub(r"(?i)^(men|women|unisex)(\'s)?\s+", "", t).strip()
        return t

    sup["_desc_title_norm"] = _series_str_clean(sup["_desc_title_norm"]).map(_clean_desc_for_display)
    sup["_title_name_raw"] = _series_str_clean(sup["_title_name_raw"]).map(_clean_desc_for_display)

    sup["_gender_title"] = _series_str_clean(sup["_gender_std"]).map(_gender_for_title)
    sup["_desc_title"] = _series_str_clean(sup["_desc_title_norm"]).map(_title_case_preserve_registered)
    sup["_color_title"] = _series_str_clean(sup["_color_in"]).map(_title_case_preserve_registered)

    # Avoid duplicating colour in Title if it is already present in the description text
    _desc_l = sup["_desc_title_norm"].astype(str).str.lower()
    _col_l = sup["_color_in"].astype(str).str.lower()
    mask_col_dup = _col_l.str.strip().ne("") & _desc_l.str.contains(_col_l.str.strip(), regex=False)
    sup.loc[mask_col_dup, "_color_title"] = ""
    base_title = (sup["_gender_title"].str.strip() + " " + sup["_desc_title"].str.strip()).str.strip()

    # Rule: Gender + Description + " - " + Color (color NON-standardized), but avoid duplicate color
    def _append_color_if_needed(bt: str, col: str) -> str:
        bt = str(bt or "").strip()
        col = str(col or "").strip()
        if not col:
            return bt
        if col.lower() in bt.lower():
            return bt
        return f"{bt} - {col}".strip()

    sup["_title"] = [
        _append_color_if_needed(bt, ct)
        for bt, ct in zip(base_title.tolist(), sup["_color_title"].astype(str).tolist())
    ]

    # Max 200 chars (truncate)
    sup["_title"] = sup["_title"].astype(str).map(lambda x: str(x)[:200].rstrip())
    # Handle: Vendor + Gender + Description + Color (color NON-standardized)
    def _make_handle(r):
        # When description is long and moved to Body (HTML), build handle from Style Name/Name (same rule as Title)
        base_text = r.get("_title_name_raw") if r.get("_desc_is_long") and r.get("_title_name_raw") else r.get("_desc_handle")
        base_text = _strip_gender_tokens(base_text)
        desc_for_handle = _strip_reg_for_handle(base_text)
        color_for_handle = _strip_reg_for_handle(r.get("_color_in", ""))
        # Avoid duplicating color if it's already present in the base text
        if color_for_handle and _norm(color_for_handle).lower() in _norm(desc_for_handle).lower():
            color_for_handle = ""
        parts = [
            _strip_reg_for_handle(r.get("_vendor")),
            _strip_reg_for_handle(r.get("_gender_std")),
            desc_for_handle,
            color_for_handle,
        ]
        parts = [p for p in parts if p and str(p).strip()]
        return slugify(" ".join(parts))
    sup["_handle"] = sup.apply(_make_handle, axis=1).apply(_remove_size_from_handle)

    # Custom Product Type: match using multiple fields (description + title + optional source product type)
    # This ensures keywords like Gilet/Bibs/Long Bibs/Bidon/Baselayer are detected even if not present in DESCRIPTION.
    sup["_product_type"] = sup["_desc_raw"].apply(lambda t: _best_match_product_type(t, product_types))

    # Optional: use a source product type column from supplier file if present
    product_type_src_col = _first_existing_col(sup, ["Product Type", "product type", "Type", "Category", "Product category", "Product Category"])
    sup["_product_type_src_raw"] = ""
    if product_type_src_col is not None:
        sup["_product_type_src_raw"] = _series_str_clean(sup[product_type_src_col])

    _pt_blob = (
        sup.get("_title", "").fillna("") + " " + sup.get("_title_name_raw", "").fillna("") + " " +
        sup["_desc_raw"].fillna("") + " " +
        sup["_product_type_src_raw"].fillna("")
    ).astype(str).str.lower()

    

    # If Custom Product Type is still empty, try matching again using a richer text blob
    # (Title + Style Name/Name + Description + optional source product type).
    sup["_product_type"] = sup["_product_type"].where(
        sup["_product_type"].astype(str).str.strip().ne(""),
        _pt_blob.apply(lambda t: _best_match_product_type(t, product_types)),
    )
# Keyword overrides (case-insensitive) – priority order matters (e.g., "long bibs" before "bibs")
    sup.loc[_pt_blob.str.contains(r"\blong\s+bibs\b", regex=True), "_product_type"] = _canon_product_type("Bib Tights")
    sup.loc[_pt_blob.str.contains(r"\bbibs\b", regex=True), "_product_type"] = _canon_product_type("Bib Shorts")
    sup.loc[_pt_blob.str.contains(r"\bgilet\b", regex=True), "_product_type"] = _canon_product_type("Vests")
    sup.loc[_pt_blob.str.contains(r"\bbidon\b", regex=True), "_product_type"] = _canon_product_type("Water Bottle")
    sup.loc[_pt_blob.str.contains(r"\bbaselayer\b", regex=True), "_product_type"] = _canon_product_type("Base Layer")
    sup.loc[_pt_blob.str.contains(r"\bt[-\s]?shirt\b", regex=True), "_product_type"] = _canon_product_type("T-Shirts")
    sup.loc[_pt_blob.str.contains(r"\btee\b", regex=True), "_product_type"] = _canon_product_type("T-Shirts")
    # Final enforcement: always output canonical Product Types from Help Data
    sup["_product_type"] = sup["_product_type"].apply(_canon_product_type)



    # -----------------------------------------------------
    # Product type gendering (Help Data -> Product Types)
    # -----------------------------------------------------
    # Default behavior is considered "Genré" unless explicitly marked NON Genré.
    sup["_is_gendered"] = sup["_product_type"].apply(
        lambda pt: product_type_gendered_map.get(str(pt or "").strip().lower(), True) if str(pt or "").strip() else True
    )

    # Gender to export:
    # 1) NON Genré -> blank
    # 2) Genré -> keep existing rule (_gender_std)
    # 3) Genré but empty -> default to "Men"
    def _gender_final(r) -> str:
        if not bool(r.get("_is_gendered", True)):
            return ""
        g = _norm(r.get("_gender_std", ""))
        return g if g else "Men"

    sup["_gender_final"] = sup.apply(_gender_final, axis=1)

    # Tags (keep standardized color/gender tags)
    # -----------------------------------------------------
    # Seasonality key (to apply Seasonality Tags per style)
    # -----------------------------------------------------
    style_num_col = _first_existing_col(sup, ["Style Number", "Style Num", "Style #", "style number", "style #", "Style"])
    style_name_col = _first_existing_col(sup, ["Style Name", "style name", "Product Name", "Name"])
    sup["_seasonality_key"] = ""
    if style_num_col is not None:
        sup["_seasonality_key"] = _series_str_clean(sup[style_num_col]).map(_clean_style_key)
    elif style_name_col is not None:
        sup["_seasonality_key"] = _series_str_clean(sup[style_name_col]).map(_clean_style_key)

    def _make_tags(r):
        tags = []
        if r["_vendor"]:
            tags.append(r["_vendor"])
        if r["_color_std"]:
            tags.append(r["_color_std"])
        # Colour-based tags
        if _norm(r["_color_std"]).lower() == "black":
            tags.append("Core")
        else:
            tags.append("Seasonal")

        if r.get("_is_gendered", True) and r.get("_gender_final", ""):
            tags.append(r["_gender_final"])
        tags.append("_badge_new")
        if r["_product_type"]:
            tags.append(r["_product_type"])
        # Event/Promotion Related (applies to entire file)
        if event_promo_tag:
            tags.append(event_promo_tag)

        # Seasonality tag (per style)
        stg = style_season_map.get(_clean_style_key(r.get("_seasonality_key", "")))
        if stg:
            tags.append(stg)

        return ", ".join([t for t in tags if t])

    sup["_tags"] = sup.apply(_make_tags, axis=1)

    # SKU
    sup["_external_id"] = _series_str_clean(sup[extid_col]).map(_norm) if extid_col else ""
    sup["_product_code"] = _series_str_clean(sup[product_col]).map(_norm) if product_col else ""

    # Variant SKU
    # -----------------------------------------------------
    # Règle par marque:
    #   a) Satisfy: style code + "-" + Size
    #   b) Norda: Style Number + "-" + Size
    #   c) Café du Cycliste: SKU + "-" + Size
    #
    # Sinon (autres fournisseurs): prendre la 1ère colonne non vide dans cet ordre: SKU, SKU 1, SKU1
    # Si aucune donnée: laisser vide (et la règle de surlignage jaune existante s'applique).
    sku_col = _first_existing_col(sup, ["SKU", "SKU 1", "SKU1", "sku", "sku 1", "sku1"])
    sku1_col = _first_existing_col(sup, ["SKU 1", "sku 1"])
    sku1_nospace_col = _first_existing_col(sup, ["SKU1", "sku1"])

    style_code_col = _first_existing_col(
        sup,
        ["Style Code", "Style code", "STYLE CODE", "Style ID", "Style", "Style Number", "Style No", "Style #", "Style#"],
    )
    style_number_col = _first_existing_col(
        sup,
        ["Style Number", "Style Num", "Style #", "Style#", "style number", "style #", "Style No", "Style"],
    )

    sup["_style_code_sku"] = _series_str_clean(sup[style_code_col]).map(_norm) if style_code_col else ""
    sup["_style_number_sku"] = _series_str_clean(sup[style_number_col]).map(_norm) if style_number_col else ""
    sup["_sku_fallback"] = ""
    if sku_col is not None:
        sup["_sku_fallback"] = _series_str_clean(sup[sku_col]).map(_norm)
    if sku1_col is not None:
        s2 = _series_str_clean(sup[sku1_col]).map(_norm)
        sup["_sku_fallback"] = sup["_sku_fallback"].where(sup["_sku_fallback"].ne(""), s2)
    if sku1_nospace_col is not None:
        s3 = _series_str_clean(sup[sku1_nospace_col]).map(_norm)
        sup["_sku_fallback"] = sup["_sku_fallback"].where(sup["_sku_fallback"].ne(""), s3)

    def _clean_hyphens(s: str) -> str:
        return re.sub(r"\s*-\s*", "-", _norm(s))

    def _make_sku(r):
        brand_key = _norm_key(r.get("_brand_choice", "")) or _norm_key(r.get("_vendor", ""))
        size = _clean_hyphens(r.get("_opt1_value", ""))

        if brand_key == "satisfy":
            base = _clean_hyphens(r.get("_style_code_sku", ""))
            if base and size:
                return f"{base}-{size}"
            return ""

        if brand_key == "norda":
            base = _clean_hyphens(r.get("_style_number_sku", ""))
            if base and size:
                return f"{base}-{size}"
            return ""

        if brand_key == "cafe du cycliste":
            base = _clean_hyphens(r.get("_sku_fallback", ""))
            if base and size:
                return f"{base}-{size}"
            return ""

        # Autres fournisseurs: SKU, SKU 1, SKU1 (dans cet ordre, selon le 1er rencontré)
        base = _clean_hyphens(r.get("_sku_fallback", ""))
        if base:
            return base
        return ""

    sup["_variant_sku"] = sup.apply(_make_sku, axis=1)


    # Barcode
    # Variant Barcode: conserver la logique actuelle (UPC puis EAN),
    # et ajouter GTIN puis GTIN 1 (dans cet ordre) comme fallbacks supplémentaires.
    gtin_col = _first_existing_col(sup, ["GTIN", "gtin"])
    gtin1_col = _first_existing_col(sup, ["GTIN 1", "GTIN1", "gtin 1", "gtin1"])

    sup["_barcode"] = sup[upc_col].apply(_barcode_keep_zeros) if upc_col else ""
    if ean_col:
        ean_series = sup[ean_col].apply(_barcode_keep_zeros)
        sup["_barcode"] = sup["_barcode"].where(sup["_barcode"].astype(str).str.strip().ne(""), ean_series)

    if gtin_col is not None:
        gtin_series = sup[gtin_col].apply(_barcode_keep_zeros)
        sup["_barcode"] = sup["_barcode"].where(sup["_barcode"].astype(str).str.strip().ne(""), gtin_series)

    if gtin1_col is not None:
        gtin1_series = sup[gtin1_col].apply(_barcode_keep_zeros)
        sup["_barcode"] = sup["_barcode"].where(sup["_barcode"].astype(str).str.strip().ne(""), gtin1_series)
    # Country (standardize)
    sup["_origin_raw"] = _series_str_clean(sup[origin_col]).map(_strip_made_in) if origin_col else ""
    sup["_origin_std"] = sup["_origin_raw"].apply(lambda x: _standardize_country(x, country_code_map))

    # HS Code
    sup["_hs"] = sup[hs_col].apply(_hs_code_clean) if hs_col else ""

    # Grams
    if grams_col:
        sup["_grams"] = _series_str_clean(sup[grams_col]).map(_norm)
    else:
        # Fallback: use Help Data -> "Variant Weight (Grams)" mapped by Custom Product Type
        sup["_grams"] = sup["_product_type"].apply(lambda pt: variant_weight_map.get(str(pt).strip().lower(), "") if pt else "")

    # Price
    if detected_price_col is not None and _header_has_cad(detected_price_col):
        # Standard CAD column: parse numeric and apply psychological rounding
        price_num = pd.to_numeric(
            sup[detected_price_col].astype(str).str.replace("$", "", regex=False).str.replace(",", "", regex=False),
            errors="coerce",
        )
        sup["_price"] = price_num.apply(_round_to_nearest_9_99)
    else:
        # Vendor-specific override: SATISFY often provides mixed currencies in the same column.
        if vendor_key in ("satisfy",):
            # Try to find a usable price column even if header doesn't mention CAD
            satisfy_price_col = _find_col(sup.columns, [
                "Retail CAD", "Retail (CAD)", "CAD Retail", "RetailCAD", "retail cad",
                "Retail Price (CAD)", "Cad MSRP", "MSRP", "msrp",
                "Retail", "retail", "Price", "price", "Retail Price", "retail price",
            ]) or detected_price_col

            if satisfy_price_col:
                price_raw = _series_str_clean(sup[satisfy_price_col]).str.strip()

                # Blank when EUR/€ is present
                is_eur = price_raw.str.contains(r"(?i)\bEUR\b|€", regex=True)

                # If any currency marker exists and it is not CAD -> blank
                has_currency = price_raw.str.contains(r"(?i)\b(?:CAD|EUR|USD)\b|€|\$", regex=True)
                is_cad = price_raw.str.contains(r"(?i)\bCAD\b", regex=True)
                reject = is_eur | (has_currency & ~is_cad)

                # Parse numeric portion
                num = price_raw.str.replace("$", "", regex=False).str.replace(",", "", regex=False).str.extract(r"([-+]?\d*\.?\d+)")[0]
                price_num = pd.to_numeric(num, errors="coerce")

                rounded = price_num.apply(_round_to_nearest_9_99)

                # Blank when rejected or non-positive (no negative/zero prices)
                rounded = rounded.where(~reject, other=float("nan"))
                rounded = rounded.where(rounded > 0, other=float("nan"))

                # Convert NaN to "" for Shopify export
                sup["_price"] = rounded.apply(lambda x: "" if (x is None or (isinstance(x, float) and math.isnan(x))) else x)
            else:
                sup["_price"] = ""
        else:
            sup["_price"] = ""
    # Cost (leave blank unless CAD column detected per rules)
    if vendor_key in ("satisfy",) and detected_cost_col is not None:
        raw = _series_str_clean(sup[detected_cost_col]).str.strip()

        # Reject EUR/€ and any explicit currency marker that is not CAD
        is_eur = raw.str.contains(r"(?i)\bEUR\b|€", regex=True)
        has_currency = raw.str.contains(r"(?i)\b(?:CAD|EUR|USD)\b|€|\$", regex=True)
        is_cad = raw.str.contains(r"(?i)\bCAD\b", regex=True)
        reject = is_eur | (has_currency & ~is_cad)

        num = raw.str.replace("$", "", regex=False).str.replace(",", "", regex=False).str.extract(r"([-+]?\d*\.?\d+)")[0]
        cost_num = pd.to_numeric(num, errors="coerce")

        # Blank if rejected or non-positive
        cost_num = cost_num.where(~reject, other=float("nan"))
        cost_num = cost_num.where(cost_num > 0, other=float("nan"))

        sup["_cost"] = cost_num.apply(lambda x: "" if (x is None or (isinstance(x, float) and math.isnan(x))) else x)
    else:
        if detected_cost_col is not None and _header_has_cad(detected_cost_col):
            sup["_cost"] = _series_str_clean(sup[detected_cost_col]).map(_norm)
        else:
            sup["_cost"] = ""

    # Size comment
    def _size_comment(r):
        key = (r["_brand_choice"] or r["_vendor"]).strip().lower()
        return size_comment_map.get(key, "")

    sup["_size_comment"] = sup.apply(_size_comment, axis=1)

    # Categories: match using DESCRIPTION (to catch LONG SLEEVE, TEE → tshirt)
    sup["_shopify_cat_id"] = sup["_desc_raw"].apply(lambda t: _best_match_id(t, shopify_cat_rows))
    sup["_google_cat_id"] = sup["_desc_raw"].apply(lambda t: _best_match_id(t, google_cat_rows))

    # Siblings
    sup["_siblings"] = sup["_handle"]

            # SEO Title & SEO Description rules (aligned with Title rules)
    # 1) Vendor + Gender ('s if Men/Women) + Description + " - " + Color (NON-standardized)
    # 2) Title Case, preserving ® ™ and TM
    # 3) Max 200 chars
    # 4) If original supplier Description > 200 chars (moved to Body), use Style Name/Name for Description part
        # SEO Title: aligned with Title rules
    # Vendor + Gender ('s if Men/Women) + Description/Style Name + " - " + Color (NON-standardized)
    def _seo_base(r) -> str:
        vendor = _title_case_preserve_registered(_norm(r.get("_vendor", "")))

        g = _norm(r.get("_gender_std", ""))
        # SEO Title: same gender rule as Title (ONLY Women's; never Men/Unisex)
        gl = g.lower().replace("’", "'") if g else ""
        if gl in ("women", "womens", "women's", "female", "femme", "femmes"):
            g = "Women's"
        else:
            g = ""
        g = _title_case_preserve_registered(g)

        # Description part: swap to Style Name/Name when source description is long
        desc_src = r.get("_title_name_raw") if r.get("_desc_is_long") and r.get("_title_name_raw") else r.get("_desc_seo", "")
        desc_src = _clean_desc_for_display(desc_src)
        desc_part = _title_case_preserve_registered(desc_src)

        # Color NON-standardized, avoid duplicates and gender markers
        color_src = _norm(r.get("_color_in", ""))
        if color_src and color_src.lower() in desc_src.lower():
            color_src = ""
        color_part = _title_case_preserve_registered(color_src)

        base = " ".join([p for p in [vendor, g, desc_part] if p]).strip()
        if color_part:
            base = f"{base} - {color_part}".strip()

        return str(base)[:200].rstrip()

    sup["_seo_title"] = sup.apply(_seo_base, axis=1)

    sup["_seo_title"] = sup["_seo_title"].apply(_scrub_nan_token_in_title)
    
    sup["_seo_title"] = sup["_seo_title"].apply(_strip_size_tokens)
# SEO Description: RESTORE previous behavior
    # Prefix fixe + contenu marque (help data -> SEO Description Brand Part), sinon fallback générique
    def _seo_desc(r):
        prefix = f"Shop the {r['_seo_title']} with free worldwide shipping, and 30-day returns on leclub.cc. "
        brand_name = _norm(r.get("_brand_choice") or r.get("_vendor"))
        brand_disp = _title_case_preserve_registered(brand_name)

        bkey = brand_name.strip().lower()
        if bkey and bkey in brand_desc_map:
            part = _norm(brand_desc_map[bkey]).rstrip().rstrip(".")
            return f"{prefix}Discover {brand_disp} {part}."
        return f"{prefix}Discover {brand_disp} products."

    sup["_seo_desc"] = sup.apply(_seo_desc, axis=1)

# behind the brand

    def _behind_brand(r):
        bkey = (r["_brand_choice"] or "").strip().lower()
        return brand_lines_map.get(bkey, "") if bkey else ""

    sup["_behind_the_brand"] = sup.apply(_behind_brand, axis=1)

    # ---------------------------------------------------------
    # Composition -> Metafield: my_fields.product_features
    # ---------------------------------------------------------
    # If a column named 'composition' (any case) exists, map it to product_features.
    composition_col = None
    for c in list(sup.columns):
        if _colkey(c) == "composition":
            composition_col = c
            break
    if composition_col is not None:
        sup["_product_features"] = _series_str_clean(sup[composition_col]).map(_sanitize_text_like_html)
    else:
        sup["_product_features"] = ""


    
    # ---------------------------------------------------------
    # v12 Option1 rules
    # ---------------------------------------------------------
    # Default behavior: Option1 = Size (with gender prefix stripped)
    sup["_opt1_name"] = "Size"
    sup["_opt1_value"] = sup["_size_std"].map(_strip_gender_prefix_size)

    # Rule 2: One Size (OS / One Size / variants) -> Title / Default Title
    mask_onesize = sup["_size_std"].astype(str).apply(_is_onesize)
    sup.loc[mask_onesize, "_opt1_name"] = "Title"
    sup.loc[mask_onesize, "_opt1_value"] = "Default Title"

    # Rule 3: If a style has only ONE row in the supplier file -> Title / Default Title
    style_num_col_v12 = _first_existing_col(sup, ["Style Number", "Style Num", "Style #", "style number", "style #", "Style NO", "Style No", "STYLE NO", "style no"])
    style_name_col_v12 = _first_existing_col(sup, ["Style Name", "style name", "STYLE NAME", "Product Name", "Name"])
    sup["_style_key_v12"] = ""
    if style_num_col_v12 is not None:
        sup["_style_key_v12"] = _series_str_clean(sup[style_num_col_v12]).map(_clean_style_key)
    elif style_name_col_v12 is not None:
        sup["_style_key_v12"] = _series_str_clean(sup[style_name_col_v12]).map(_clean_style_key)

    _k = sup["_style_key_v12"].astype(str).str.strip()
    counts = _k[_k.ne("")].value_counts()
    mask_single_style = _k.map(lambda x: counts.get(x, 0) == 1 if x else False)
    # Ne pas forcer "Default Title" si une vraie taille est présente (ex: XS, S, M, etc.)
    _size_clean = sup["_size_std"].map(_strip_gender_prefix_size).astype(str).str.strip()
    _dash_tokens = {"-", "–", "—"}
    mask_has_real_size = _size_clean.ne("") & (~sup["_size_std"].astype(str).apply(_is_onesize)) & (~_size_clean.isin(_dash_tokens))
    mask_single_style = mask_single_style & (~mask_has_real_size)
    sup.loc[mask_single_style, "_opt1_name"] = "Title"
    sup.loc[mask_single_style, "_opt1_value"] = "Default Title"

# ---------------------------------------------------------
    # ---------------------------------------------------------
    # Variant SKU + MPN rules (after Option1 is finalized)
    # ---------------------------------------------------------
    def _clean_hyphens_sku(s: str) -> str:
        # remove spaces around hyphens and normalize whitespace
        return re.sub(r"\s*-\s*", "-", _norm(s))

    def _make_variant_sku(r):
        # Identify brand/vendor (case-insensitive)
        brand_key = _norm_key(r.get("_brand_choice", "")) or _norm_key(r.get("_vendor", ""))
                # Size doit provenir de 'Variant Metafield: mm-google-shopping.size' (via _size_std)
        size_raw = _strip_gender_prefix_size(r.get("_size_std", ""))
        size = _clean_hyphens_sku(size_raw)
        # Ne jamais utiliser 'Default Title' comme taille
        if _norm_key(size) in ("default title", "default", "default value"):
            size = ""

        if brand_key == "satisfy":
            base = _clean_hyphens_sku(r.get("_style_code_sku", ""))
            return f"{base}-{size}" if base and size else ""

        if brand_key == "norda":
            base = _clean_hyphens_sku(r.get("_style_number_sku", ""))
            return f"{base}-{size}" if base and size else ""

        if brand_key in ("cafe du cycliste", "café du cycliste"):
            base = _clean_hyphens_sku(r.get("_sku_fallback", ""))
            return f"{base}-{size}" if base and size else ""

        # Other suppliers: first non-empty among SKU / SKU 1 / SKU1 (already resolved into _sku_fallback)
        base = _clean_hyphens_sku(r.get("_sku_fallback", ""))
        return base if base else ""

    sup["_variant_sku"] = sup.apply(_make_variant_sku, axis=1)

    # Build output (strict order)
    # ---------------------------------------------------------
    out = pd.DataFrame(columns=SHOPIFY_OUTPUT_COLUMNS)

    out["Handle"] = sup["_handle"]
    out["Command"] = "NEW"
    out["Title"] = sup["_title"]
    out["Body (HTML)"] = sup["_body_html"]
    out["Vendor"] = sup["_vendor"]
    out["Custom Product Type"] = sup["_product_type"]
    out["Tags"] = sup["_tags"]

    out["Published"] = False
    out["Published Scope"] = "global"

    out["Option1 Name"] = sup["_opt1_name"]
    out["Option1 Value"] = sup["_opt1_value"]

    out["Variant SKU"] = sup["_variant_sku"]
    out["Variant Barcode"] = sup["_barcode"]
    out["Variant Country of Origin"] = sup["_origin_std"]
    out["Variant HS Code"] = sup["_hs"]
    out["Variant Grams"] = sup["_grams"]

    out["Variant Inventory Tracker"] = "shopify"
    out["Variant Inventory Policy"] = "deny"
    out["Variant Fulfillment Service"] = "manual"
    out["Variant Price"] = sup["_price"]

    out["Variant Requires Shipping"] = True
    out["Variant Taxable"] = True

    out["SEO Title"] = sup["_seo_title"]
    out["SEO Description"] = sup["_seo_desc"]

    # Final safety: ensure no <br> or &nbsp; artifacts in text fields
    out["Body (HTML)"] = out["Body (HTML)"].map(_sanitize_text_like_html)
    out["Metafield: my_fields.product_features [multi_line_text_field]"] = out["Metafield: my_fields.product_features [multi_line_text_field]"].map(_sanitize_text_like_html)

    out["Variant Weight Unit"] = "g"
    out["Cost per item"] = sup["_cost"]
    out["Status"] = "draft"

    out["Metafield: my_fields.product_use_case [multi_line_text_field]"] = ""
    out["Metafield: my_fields.product_features [multi_line_text_field]"] = sup["_product_features"]
    out["Metafield: my_fields.behind_the_brand [multi_line_text_field]"] = sup["_behind_the_brand"]
    out["Metafield: my_fields.size_comment [single_line_text_field]"] = sup["_size_comment"]
    out["Metafield: my_fields.gender [single_line_text_field]"] = sup["_gender_final"]

    out["Metafield: my_fields.colour [single_line_text_field]"] = sup["_color_std"]
    out["Metafield: mm-google-shopping.color"] = sup["_color_std"]
    out["Variant Metafield: mm-google-shopping.size"] = sup["_size_std"].map(_strip_gender_prefix_size)

    out["Metafield: mm-google-shopping.size_system"] = "US"
    out["Metafield: mm-google-shopping.condition"] = "new"
    out["Metafield: mm-google-shopping.google_product_category"] = sup["_google_cat_id"]
    out["Metafield: mm-google-shopping.gender"] = sup["_gender_final"]

    out["Variant Metafield: mm-google-shopping.mpn"] = sup["_variant_sku"]
    out["Variant Metafield: mm-google-shopping.gtin"] = sup["_barcode"]

    out["Metafield: theme.siblings [single_line_text_field]"] = sup["_siblings"]
    out["Category: ID"] = sup["_shopify_cat_id"]

    out["Inventory Available: Boutique"] = 0
    out["Inventory Available: Le Club"] = 0

    out = out.reindex(columns=SHOPIFY_OUTPUT_COLUMNS)
    out = out.where(out.notna(), "")
    # Also remove stringified NaN/None that can appear after astype(str)
    out = out.replace({r"^\s*(nan|none)\s*$": ""}, regex=True)
    # Remove any remaining embedded "nan" tokens (e.g., "nan - Fireclay") that can slip in via concatenation
    out = out.replace({r"(?i)^\s*nan\s*-\s*": "", r"(?i)\bnan\b": ""}, regex=True)
    out = out.replace({r"\s{2,}": " "}, regex=True)  # éviter "nan" dans l'export

    # Internal flag for styling (not exported)
    out["OUT_COLOR_HIT"] = sup.get("_color_map_hit", True)


    # Yellow rules
    yellow_if_empty_cols = [
        "Handle",
        "Title",
        "Body (HTML)",
        "Custom Product Type",
        "Option1 Name",
        "Option1 Value",
        "Variant Price",
        "Variant Grams",
        "Variant Country of Origin",
        "Variant HS Code",
        "SEO Title",
        "SEO Description",
        "Metafield: my_fields.size_comment [single_line_text_field]",
        "Metafield: my_fields.colour [single_line_text_field]",
        "Metafield: mm-google-shopping.color",
        "Variant Metafield: mm-google-shopping.size",
        "Metafield: mm-google-shopping.google_product_category",
        "Category: ID",
        "Variant SKU",
        "Variant Metafield: mm-google-shopping.mpn",
    ]

    def _apply_red_font_for_tags(buffer: io.BytesIO, sheet_name: str, rows_to_color_red: list[int]) -> io.BytesIO:
        """Apply red font to the 'Tags' cell for given 0-based dataframe row indexes."""
        buffer.seek(0)
        wb = openpyxl.load_workbook(buffer)
        if sheet_name not in wb.sheetnames:
            return buffer
        ws = wb[sheet_name]

        # Find Tags column index from header row (row 1)
        tags_col_idx = None
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            if str(v).strip() == "Tags":
                tags_col_idx = c
                break
        if tags_col_idx is None:
            return buffer

        red_font = openpyxl.styles.Font(color="FFFF0000")

        # Data rows start at Excel row 2
        if not rows_to_color_red:
            rows_to_color_red = []
            for excel_row in range(2, ws.max_row + 1):
                v = ws.cell(row=excel_row, column=tags_col_idx).value
                if v and "seasonal" in str(v).lower():
                    rows_to_color_red.append(excel_row - 2)

        for df_i in rows_to_color_red:
            excel_row = df_i + 2
            cell = ws.cell(row=excel_row, column=tags_col_idx)
            cell.font = red_font

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out
    def _apply_red_font_for_color_multi(buffer: io.BytesIO, sheet_name: str, cols: list[str]) -> io.BytesIO:
        """Apply red font to specified columns when the cell contains '/' (ex: multi-colour)."""
        buffer.seek(0)
        wb = openpyxl.load_workbook(buffer)
        if sheet_name not in wb.sheetnames:
            return buffer
        ws = wb[sheet_name]

        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        col_index = {str(v).strip(): i + 1 for i, v in enumerate(header) if v is not None}

        red_font = openpyxl.styles.Font(color="FFFF0000")

        for col_name in cols:
            if col_name not in col_index:
                continue
            cidx = col_index[col_name]
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=cidx)
                v = cell.value
                if v is None:
                    continue
                if "/" in str(v):
                    cell.font = red_font

        outb = io.BytesIO()
        wb.save(outb)
        outb.seek(0)
        return outb


    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:

        # Split into "products" and "do not import" based on existing Shopify file (if provided)
        existing_handles_set, existing_key_sets = _build_existing_shopify_index(existing_shopify_xlsx_bytes)

        # columns expected in output
        vendor_col = "Vendor" if "Vendor" in out.columns else None
        sku_col = "Variant SKU" if "Variant SKU" in out.columns else ("SKU" if "SKU" in out.columns else None)
        upc_col = "Variant Barcode" if "Variant Barcode" in out.columns else ("Barcode" if "Barcode" in out.columns else ("UPC" if "UPC" in out.columns else None))

        def _getcol(r, c):
            return r.get(c, "") if c else ""
        handle_col_out = "Handle" if "Handle" in out.columns else None

        mask_existing = []
        for _, r in out.iterrows():
            vendor = _getcol(r, vendor_col) or vendor_name
            sku = _getcol(r, sku_col)
            upc = _getcol(r, upc_col)
            handle_val = _getcol(r, handle_col_out)
            handle_norm = _norm_handle(handle_val) if handle_col_out else ""
            is_existing = (handle_norm in existing_handles_set) if handle_norm else False
            if not is_existing:
                is_existing = _row_is_existing(str(vendor), str(sku), str(upc), existing_key_sets)
            mask_existing.append(is_existing)

        mask_existing = pd.Series(mask_existing, index=out.index)

        products_df = out.loc[~mask_existing].copy()
        do_not_import_df = out.loc[mask_existing].copy()

        products_df[SHOPIFY_OUTPUT_COLUMNS].to_excel(writer, index=False, sheet_name="products")
        do_not_import_df[SHOPIFY_OUTPUT_COLUMNS].to_excel(writer, index=False, sheet_name="do not import")
        pd.DataFrame(warnings).to_excel(writer, index=False, sheet_name="warnings")

    
    # Red font for Tags when colour is NOT Black (i.e., Seasonal) — apply on both sheets
    existing_handles_set, existing_key_sets = _build_existing_shopify_index(existing_shopify_xlsx_bytes)

    def _rows_to_color_for_df(df_slice: pd.DataFrame) -> list[int]:
        if "_color_std" not in df_slice.columns:
            return []
        return [
            i
            for i, c in enumerate(df_slice["_color_std"].astype(str).tolist())
            if _norm(c) != "" and _norm(c).lower() != "black"
        ]

    # For handle red: when output handle already exists in Shopify
    def _rows_handle_conflict(df_slice: pd.DataFrame) -> list[int]:
        """Rows where Handle conflicts with an existing Shopify handle."""
        if "Handle" not in df_slice.columns:
            return []
        handles_norm = df_slice["Handle"].apply(_norm_handle)
        mask = handles_norm.isin(existing_handles_set) & handles_norm.ne("")
        return [i for i, v in enumerate(mask.tolist()) if v]
        return [i for i, h in enumerate(df_slice["Handle"].astype(str).tolist()) if _norm(h) in existing_handles_set and _norm(h) != ""]

    # Reload workbook buffer as BytesIO for styling helpers
    buffer.seek(0)

    # Apply tag red and yellow empty on each sheet
    buffer = _apply_red_font_for_tags(buffer, "products", _rows_to_color_for_df(products_df))
    buffer = _apply_red_font_for_tags(buffer, "do not import", _rows_to_color_for_df(do_not_import_df))

    buffer = _apply_yellow_for_empty(buffer, "products", yellow_if_empty_cols)
    buffer = _apply_yellow_for_empty(buffer, "do not import", yellow_if_empty_cols)
    # Red font for Title when it contains "?" or "/" (needs manual review)
    title_warn_cols = ["Title", "SEO Title"]

    def _rows_title_warn(df_slice: pd.DataFrame) -> list[int]:
        if "Title" not in df_slice.columns:
            return []
        s = _series_str_clean(df_slice["Title"])
        mask = s.str.contains(r"[\?/]", regex=True)
        return [i for i, v in enumerate(mask.tolist()) if v]

    buffer = _apply_red_font_for_rows_cols(buffer, "products", _rows_title_warn(products_df), title_warn_cols)
    buffer = _apply_red_font_for_rows_cols(buffer, "do not import", _rows_title_warn(do_not_import_df), title_warn_cols)

    # Red font for colour metafields when supplier colour was NOT found in Help Data mapping
    color_unmapped_cols = [
        "Metafield: my_fields.colour [single_line_text_field]",
        "Metafield: mm-google-shopping.color",
    ]

    def _rows_unmapped(df_slice: pd.DataFrame) -> list[int]:
        if "OUT_COLOR_HIT" not in df_slice.columns:
            return []
        mask = ~df_slice["OUT_COLOR_HIT"].astype(bool)
        return [i for i, v in enumerate(mask.tolist()) if v]

    buffer = _apply_red_font_for_rows_cols(buffer, "products", _rows_unmapped(products_df), color_unmapped_cols)
    buffer = _apply_red_font_for_rows_cols(buffer, "do not import", _rows_unmapped(do_not_import_df), color_unmapped_cols)

    # Red font for size metafield when supplier size was NOT found in Help Data mapping
    size_unmapped_cols = ["Variant Metafield: mm-google-shopping.size"]

    def _rows_size_unmapped(df_slice: pd.DataFrame) -> list[int]:
        if "OUT_SIZE_HIT" not in df_slice.columns:
            return []
        mask = ~df_slice["OUT_SIZE_HIT"].astype(bool)
        return [i for i, v in enumerate(mask.tolist()) if v]

    buffer = _apply_red_font_for_rows_cols(buffer, "products", _rows_size_unmapped(products_df), size_unmapped_cols)
    buffer = _apply_red_font_for_rows_cols(buffer, "do not import", _rows_size_unmapped(do_not_import_df), size_unmapped_cols)


    # Red font for multi-colour values (contains "/") on colour columns
    color_cols_multi = [
        "Metafield: my_fields.colour [single_line_text_field]",
        "Metafield: mm-google-shopping.color",
    ]
    buffer = _apply_red_font_for_color_multi(buffer, "products", color_cols_multi)
    buffer = _apply_red_font_for_color_multi(buffer, "do not import", color_cols_multi)


    # Apply red font for handle conflicts (only the cell in Handle column)
    buffer = _apply_red_font_for_handle(buffer, "products", _rows_handle_conflict(products_df))
    buffer = _apply_red_font_for_handle(buffer, "do not import", _rows_handle_conflict(do_not_import_df))
    # Header notes (Excel comments) to explain red formatting / validations
    header_notes = {
        "Handle": "ROUGE = Le handle existe déjà dans le fichier d’inventaire fourni.",
        "Title": "ROUGE = Le titre comporte un des deux caractères suivants: ? ou /.",
        "SEO Title": "ROUGE = Le titre comporte un des deux caractères suivants: ? ou /.",
        "Tags": "ROUGE = Assurez-vous que les tags Seasonal sont bien les bons.",
        "Metafield: my_fields.colour [single_line_text_field]": "ROUGE = Les couleurs ne sont pas présentes dans le mapping (Help Data).",
        "Metafield: mm-google-shopping.color": "ROUGE = Les couleurs ne sont pas présentes dans le mapping (Help Data).",
        "Custom Product Type": "Assurez-vous que les catégories trouvées sont bien les bonnes.",
        "Metafield: mm-google-shopping.google_product_category": "Assurez-vous que les catégories trouvées sont bien les bonnes.",
    }

    buffer = _apply_header_notes(buffer, "products", header_notes)
    buffer = _apply_header_notes(buffer, "do not import", header_notes)


    return buffer.getvalue(), pd.DataFrame(warnings)
